from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.contrib.auth.models import User
from django.http import HttpResponse
from .models import Bill, Payment
from .serializers import UserSerializer, BillSerializer, PaymentSerializer
import openpyxl
from decimal import Decimal
from datetime import datetime

@api_view(['GET'])
def api_root(request):
    return Response({
        "message": "Welcome to the DELA Water Billing System API!",
        "endpoints": {
            "admin_bills": "/api/admin/bills/<int:pk>/",
            "user_bills": "/api/user/bills/",
            "user_profile": "/api/user/profile/",
            "register": "/api/register/",
            "token_obtain": "/api/token/",
            "token_refresh": "/api/token/refresh/",
            "admin_users": "/api/admin/users/",
            "admin_allocate_bill": "/api/admin/bills/",
            "admin_export_bills": "/api/admin/export-bills/",
        }
    })

# Admin: Update bills
class BillUpdateView(generics.UpdateAPIView):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    permission_classes = [permissions.IsAdminUser]  # Only admin can update bills

# User: View bills
class UserBillListView(generics.ListAPIView):
    serializer_class = BillSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Bill.objects.filter(customer=self.request.user)

# User: Update profile
class UserProfileUpdateView(generics.RetrieveUpdateAPIView):  # Supports GET and PUT
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user

# User: Register
class UserRegistrationView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class PaymentCreateView(generics.CreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

# Admin: Update bills
class BillUpdateView(generics.UpdateAPIView):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    permission_classes = [permissions.IsAdminUser]  # Only admin can update bills

# User: View bills
class UserBillListView(generics.ListAPIView):
    serializer_class = BillSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Bill.objects.filter(customer=self.request.user)

# User: Update profile
class UserProfileUpdateView(generics.RetrieveUpdateAPIView):  # Supports GET and PUT
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user

# User: Register
class UserRegistrationView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

# Payment: Create a payment
class PaymentCreateView(generics.CreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

# -------------------------
# New Admin Endpoints
# -------------------------

# 1. Admin: List all users
class AdminUserListView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]

# 2. Admin: Allocate (create) a new bill
class AdminBillCreateView(generics.CreateAPIView):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    permission_classes = [permissions.IsAdminUser]

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        # Expect data to include: customer (user id), total_amount, and due_date.
        try:
            total_amount = Decimal(data.get('total_amount'))
        except Exception:
            return Response({"error": "Invalid total_amount format."}, status=status.HTTP_400_BAD_REQUEST)
        data['total_amount'] = total_amount
        data['remaining_balance'] = total_amount  # Set remaining balance equal to total_amount.
        # Optionally, set billing_date to today if not provided.
        if not data.get('billing_date'):
            data['billing_date'] = datetime.now().date()
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

# 3. Admin: Export bills as an Excel file
@api_view(['GET'])
@permission_classes([permissions.IsAdminUser])
def admin_export_bills_view(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bills Report"

    # Define header row
    columns = ["User", "Total Amount", "Remaining Balance", "Billing Date", "Due Date", "Status"]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Populate rows with Bill data
    bills = Bill.objects.all().order_by('due_date')
    for bill in bills:
        row_num += 1
        row = [
            bill.customer.username,
            float(bill.total_amount),
            float(bill.remaining_balance),
            bill.billing_date.strftime("%Y-%m-%d"),
            bill.due_date.strftime("%Y-%m-%d"),
            bill.status,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bills_report.xlsx"'
    workbook.save(response)
    return response